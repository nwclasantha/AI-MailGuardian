"""
Enterprise Excel Report Generator
Beautiful, formatted Excel reports with charts and professional styling
"""

import logging
from typing import Dict, List
from datetime import datetime

logger = logging.getLogger(__name__)


class EnterpriseExcelReportGenerator:
    """Generate enterprise-level Excel reports with professional formatting"""

    def __init__(self, config, analyzer=None):
        self.config = config
        self.analyzer = analyzer  # Real analyzer instance with ML engine

    def generate_enterprise_excel(self, results: List[Dict], filename: str):
        """Generate beautifully formatted enterprise Excel report"""
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import PieChart, BarChart, Reference
            from openpyxl.utils import get_column_letter

            # Calculate statistics
            stats = self._calculate_stats(results)

            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Sheet 1: Executive Dashboard (ALWAYS has data)
                logger.info("Creating Executive Dashboard sheet...")
                self._create_executive_dashboard(writer, stats)

                # Sheet 2: Risk Analysis (ALWAYS has data)
                logger.info("Creating Risk Analysis sheet...")
                self._create_risk_analysis_sheet(writer, results, stats)

                # Sheet 3: Breach Intelligence
                logger.info("Creating Breach Intelligence sheet...")
                self._create_breach_intelligence_sheet(writer, results)

                # Sheet 4: Detailed Results (ALWAYS has data)
                logger.info("Creating Detailed Results sheet...")
                self._create_detailed_results_sheet(writer, results)

                # Sheet 5: MITRE ATT&CK
                logger.info("Creating MITRE ATT&CK sheet...")
                self._create_mitre_sheet(writer, results)

                # Sheet 6: Mitigation Actions
                logger.info("Creating Mitigation Actions sheet...")
                self._create_mitigation_sheet(writer, results)

                # Sheet 7: DNS Security
                logger.info("Creating DNS Security sheet...")
                self._create_dns_security_sheet(writer, results)

                # Sheet 8: Threat Details
                logger.info("Creating Threat Details sheet...")
                self._create_threat_details_sheet(writer, results)

                # Sheet 9: Disposable & Typosquatting Detection
                logger.info("Creating Disposable & Typosquatting sheet...")
                self._create_disposable_typosquat_sheet(writer, results)

                # Sheet 10: Vulnerability Assessment
                logger.info("Creating Vulnerability Assessment sheet...")
                self._create_vulnerability_sheet(writer, results)

                # Sheet 11: Data Exposure Summary
                logger.info("Creating Data Exposure Summary sheet...")
                self._create_data_exposure_sheet(writer, results)

                # Sheet 12: Advanced Security Checks
                logger.info("Creating Advanced Security Checks sheet...")
                self._create_advanced_security_sheet(writer, results)

                # Sheet 13: ML Prediction Statistics
                logger.info("Creating ML Prediction Statistics sheet...")
                self._create_ml_metrics_sheet(writer)

                # Sheet 13: ML Performance Charts (if ML data available)
                if self.analyzer and hasattr(self.analyzer, 'ml_engine'):
                    logger.info("Creating ML Performance Charts sheet...")
                    self._create_ml_charts_sheet(writer, results)

                # Ensure all sheets are visible BEFORE closing writer
                workbook = writer.book
                logger.info(f"Total sheets created: {len(workbook.worksheets)}")

                for sheet in workbook.worksheets:
                    logger.info(f"Sheet: {sheet.title}, State: {sheet.sheet_state}, Max Row: {sheet.max_row}, Max Col: {sheet.max_column}")
                    sheet.sheet_state = 'visible'

                logger.info(f"All {len(workbook.worksheets)} sheets set to visible")

            # Apply advanced formatting (non-critical - don't fail if this errors)
            try:
                self._apply_advanced_formatting(filename, stats)
                logger.info("Advanced formatting applied successfully")
            except Exception as fmt_error:
                logger.warning(f"Could not apply advanced formatting (file still created): {fmt_error}")

            logger.info(f"Enterprise Excel report generated: {filename}")
            return True

        except Exception as e:
            logger.error(f"Error generating enterprise Excel: {e}", exc_info=True)
            raise

    def _create_executive_dashboard(self, writer, stats: Dict):
        """Create executive dashboard sheet"""
        import pandas as pd

        # Dashboard data - all columns must have same length
        dashboard_data = {
            'Metric': [
                'Total Emails Analyzed',
                'Critical Risk Emails',
                'High Risk Emails',
                'Medium Risk Emails',
                'Low Risk / Safe',
                '',
                '--- BREACH INTELLIGENCE ---',
                'Emails in Data Breaches',
                'Breach Rate (%)',
                'Password Breaches',
                '',
                '--- FRAUD DETECTION ---',
                'Disposable/Temp Emails',
                'Typosquatting Domains',
                '',
                '--- RISK METRICS ---',
                'Average Risk Score',
                'Highest Risk Score',
                'Total Threats Detected',
                'Total Vulnerabilities',
                '',
                '--- INFRASTRUCTURE ---',
                'DNS Security Issues'
            ],
            'Value': [
                stats['total'],
                stats['critical'],
                stats['high'],
                stats['medium'],
                stats['low'],
                '',
                '',
                stats['breached'],
                f"{stats['breach_percentage']:.1f}%",
                stats.get('password_breaches', 0),
                '',
                '',
                stats.get('disposable_emails', 0),
                stats.get('typosquatting_domains', 0),
                '',
                '',
                stats['avg_risk'],
                stats['max_risk'],
                stats['total_threats'],
                stats.get('vulnerabilities', 0),
                '',
                '',
                stats['dns_issues']
            ],
            'Status': [
                '✅',
                '🚨' if stats['critical'] > 0 else '✅',
                '⚠️' if stats['high'] > 0 else '✅',
                '⚡' if stats['medium'] > 0 else '✅',
                '✅',
                '',
                '',
                '🚨' if stats['breached'] > 0 else '✅',
                '⚠️' if stats['breach_percentage'] > 20 else '✅',
                '🚨' if stats.get('password_breaches', 0) > 0 else '✅',
                '',
                '',
                '🗑️' if stats.get('disposable_emails', 0) > 0 else '✅',
                '🎭' if stats.get('typosquatting_domains', 0) > 0 else '✅',
                '',
                '',
                '⚠️' if stats['avg_risk'] > 40 else '✅',
                '🚨' if stats['max_risk'] > 70 else '⚠️' if stats['max_risk'] > 40 else '✅',
                '🔍',
                '⚠️' if stats.get('vulnerabilities', 0) > 0 else '✅',
                '',
                '',
                '⚠️' if stats['dns_issues'] > 0 else '✅'
            ]
        }

        df = pd.DataFrame(dashboard_data)
        df.to_excel(writer, sheet_name='Executive Dashboard', index=False)

        # Add recommendations section - all columns must have same length
        recommendations_data = {
            'Priority': ['1', '2', '3', '4', '5'],
            'Recommendation': [
                f'Immediate action required for {stats["critical"]} critical risk emails' if stats['critical'] > 0 else 'Maintain current security posture',
                f'Review and secure {stats["breached"]} breached email accounts' if stats['breached'] > 0 else 'Continue monitoring for new breaches',
                'Enable Two-Factor Authentication (2FA) on all accounts',
                'Implement password rotation policy (every 90 days)',
                'Conduct security awareness training for all users'
            ]
        }

        df_rec = pd.DataFrame(recommendations_data)
        # Place recommendations below dashboard data (dynamic row count + gap)
        rec_startrow = len(dashboard_data['Metric']) + 3
        df_rec.to_excel(writer, sheet_name='Executive Dashboard', index=False, startrow=rec_startrow)

    def _create_risk_analysis_sheet(self, writer, results: List[Dict], stats: Dict):
        """Create risk analysis sheet with detailed breakdown"""
        import pandas as pd

        # Risk distribution data
        risk_data = {
            'Risk Level': ['Critical', 'High', 'Medium', 'Low / Safe'],
            'Count': [stats['critical'], stats['high'], stats['medium'], stats['low']],
            'Percentage': [
                f"{stats['critical_percentage']:.1f}%",
                f"{stats['high_percentage']:.1f}%",
                f"{stats['medium_percentage']:.1f}%",
                f"{stats['low_percentage']:.1f}%"
            ],
            'Description': [
                'Immediate action required',
                'Attention needed soon',
                'Monitor regularly',
                'No immediate concerns'
            ]
        }

        df_risk = pd.DataFrame(risk_data)
        df_risk.to_excel(writer, sheet_name='Risk Analysis', index=False)

        # Score distribution (create buckets)
        score_ranges = ['0-20', '21-40', '41-60', '61-80', '81-100']
        score_counts = [
            sum(1 for r in results if 0 <= r.get('risk_score', 0) <= 20),
            sum(1 for r in results if 21 <= r.get('risk_score', 0) <= 40),
            sum(1 for r in results if 41 <= r.get('risk_score', 0) <= 60),
            sum(1 for r in results if 61 <= r.get('risk_score', 0) <= 80),
            sum(1 for r in results if 81 <= r.get('risk_score', 0) <= 100)
        ]

        score_dist_data = {
            'Score Range': score_ranges,
            'Count': score_counts,
            'Percentage': [f"{(c/len(results)*100):.1f}%" if len(results) > 0 else "0%" for c in score_counts]
        }

        df_scores = pd.DataFrame(score_dist_data)
        df_scores.to_excel(writer, sheet_name='Risk Analysis', index=False, startrow=8)

    def _create_breach_intelligence_sheet(self, writer, results: List[Dict]):
        """Create breach intelligence sheet"""
        import pandas as pd

        breach_data = []

        for result in results:
            breach_info = result.get('breach_info') or {}
            if breach_info.get('found'):
                # Build details list with fallback from breach names
                details_list = breach_info.get('details') or []
                if not isinstance(details_list, list):
                    details_list = []
                # Fallback: if details empty but breach names available, build minimal entries
                if not details_list and breach_info.get('breaches'):
                    breach_names = breach_info['breaches']
                    if isinstance(breach_names, list):
                        for bname in breach_names[:20]:
                            if isinstance(bname, str) and bname.strip():
                                details_list.append({'name': bname, 'breach_date': 'Unknown', 'data_classes': []})

                # Get top MITRE techniques for this email
                mitre_details = result.get('mitre_details', [])
                top_mitre = ', '.join([f"{t.get('id', 'N/A')} ({t.get('similarity', 0):.0f}%)" for t in mitre_details[:3]]) if mitre_details else 'N/A'

                for breach in details_list:
                    if isinstance(breach, list):
                        # Handle XposedOrNot nested list format [["Adobe","LinkedIn"]]
                        for breach_name in breach:
                            if isinstance(breach_name, str):
                                breach_data.append({
                                    'Email': result.get('email', 'N/A'),
                                    'Breach Name': breach_name,
                                    'Breach Date': 'Unknown',
                                    'Domain': 'N/A',
                                    'Affected Accounts': 'N/A',
                                    'Compromised Data': 'N/A',
                                    'Severity': (breach_info.get('severity') or 'N/A').upper(),
                                    'Verified': 'Unknown',
                                    'MITRE Techniques': top_mitre,
                                    'Description': 'Dark web breach (XposedOrNot)'
                                })
                    elif isinstance(breach, dict):
                        desc = breach.get('description') or 'N/A'
                        breach_data.append({
                            'Email': result.get('email', 'N/A'),
                            'Breach Name': breach.get('name', 'Unknown'),
                            'Breach Date': breach.get('breach_date', 'Unknown'),
                            'Domain': breach.get('domain', 'N/A'),
                            'Affected Accounts': breach.get('pwn_count', 'N/A'),
                            'Compromised Data': ', '.join(str(dc) for dc in (breach.get('data_classes') or [])[:10]),
                            'Severity': (breach_info.get('severity') or 'N/A').upper(),
                            'Verified': 'Yes' if breach.get('is_verified') else 'No',
                            'MITRE Techniques': top_mitre,
                            'Description': desc[:200] + '...' if len(desc) > 200 else desc
                        })

        if breach_data:
            df_breaches = pd.DataFrame(breach_data)
            df_breaches.to_excel(writer, sheet_name='Breach Intelligence', index=False)
        else:
            # Create empty sheet with headers AND a placeholder row
            empty_data = pd.DataFrame([{
                'Email': 'No breaches detected',
                'Breach Name': '-',
                'Breach Date': '-',
                'Domain': '-',
                'Status': 'Clean'
            }])
            empty_data.to_excel(writer, sheet_name='Breach Intelligence', index=False)

    def _create_detailed_results_sheet(self, writer, results: List[Dict]):
        """Create comprehensive detailed results sheet with ALL detection columns"""
        import pandas as pd

        detailed_data = []

        for result in results:
            breach_info = result.get('breach_info') or {}
            dns = result.get('dns_security') or {}
            ml_preds = result.get('ml_predictions') or {}
            threats = result.get('threats') or []
            typosquat_info = result.get('typosquat_info') or {}
            dom_rep = result.get('domain_reputation') or {}

            # Detection flags
            is_disposable = any(t.get('type') == 'disposable_email' for t in threats)
            is_typosquat = typosquat_info.get('is_typosquat', False)
            has_password_breach = (result.get('password_breach') or {}).get('found', False)

            # ML scores
            ensemble = ml_preds.get('ensemble')
            anomaly = ml_preds.get('anomaly_score')

            # Threat types
            threat_types = list(set(t.get('type', '') for t in threats))

            detailed_data.append({
                'Email': result.get('email', 'N/A'),
                'Domain': result.get('email', 'N/A').split('@')[1] if '@' in result.get('email', '') else 'N/A',
                'Risk Score': result.get('risk_score', 0),
                'Risk Level': (result.get('risk_level') or 'unknown').upper(),
                'Breached': 'YES' if breach_info.get('found') else 'NO',
                'Breach Count': breach_info.get('count', 0) if breach_info.get('found') else 0,
                'Breach Severity': (breach_info.get('severity') or 'N/A').upper() if breach_info.get('found') else 'N/A',
                'Password Breach': 'COMPROMISED' if has_password_breach else 'Safe',
                'Disposable Email': 'YES' if is_disposable else 'NO',
                'Typosquatting': typosquat_info.get('target_domain', 'NO') if is_typosquat else 'NO',
                'Typosquat Attack': (typosquat_info.get('attack_type') or '').replace('_', ' ').title() if is_typosquat else 'N/A',
                'Threats Detected': len(threats),
                'Threat Types': ', '.join(t.replace('_', ' ').title() for t in threat_types) if threat_types else 'None',
                'Vulnerabilities': len(result.get('vulnerabilities', [])),
                'Domain Age (days)': dom_rep.get('age', 'Unknown'),
                'Domain Reputation': dom_rep.get('score', 'N/A'),
                'Domain Flags': ', '.join(str(f) for f in (dom_rep.get('flags') or [])[:5]) if dom_rep.get('flags') else 'None',
                'ML Ensemble': f"{float(ensemble):.0%}" if ensemble is not None else 'N/A',
                'Anomaly Score': f"{float(anomaly):.0%}" if anomaly is not None else 'N/A',
                'DNS Score': dns.get('score', 'N/A'),
                'SPF': 'Configured' if dns.get('spf') else 'Missing',
                'DMARC': 'Configured' if dns.get('dmarc') else 'Missing',
                'DKIM': 'Configured' if dns.get('dkim') else 'Missing',
                'DNSSEC': 'Enabled' if dns.get('dnssec') else 'Disabled',
                'MITRE Techniques': ', '.join(t.get('id', '') for t in result.get('mitre_details', [])[:3]) if result.get('mitre_details') else 'None'
            })

        df_detailed = pd.DataFrame(detailed_data)
        # Sort by risk score descending
        df_detailed = df_detailed.sort_values('Risk Score', ascending=False)
        df_detailed.to_excel(writer, sheet_name='Detailed Results', index=False)

    def _create_mitre_sheet(self, writer, results: List[Dict]):
        """Create MITRE ATT&CK techniques sheet"""
        import pandas as pd

        mitre_data = []

        for result in results:
            if result.get('mitre_details'):
                for tech in result['mitre_details']:
                    mitre_data.append({
                        'Email': result.get('email', 'N/A'),
                        'Technique ID': tech.get('id', 'N/A'),
                        'Technique Name': tech.get('name', 'N/A'),
                        'Tactic': tech.get('tactic', 'N/A'),
                        'Confidence': f"{tech.get('similarity', 0):.1f}%",
                        'Severity': tech.get('severity', 'N/A'),
                        'Description': (lambda d: d[:300] + '...' if len(d) > 300 else d)(tech.get('description') or 'N/A')
                    })

        if mitre_data:
            df_mitre = pd.DataFrame(mitre_data)
            df_mitre.to_excel(writer, sheet_name='MITRE ATT&CK', index=False)
        else:
            # Create placeholder row
            empty_data = pd.DataFrame([{
                'Email': 'No MITRE techniques detected',
                'Technique ID': '-',
                'Technique Name': '-',
                'Tactic': '-',
                'Confidence': '-'
            }])
            empty_data.to_excel(writer, sheet_name='MITRE ATT&CK', index=False)

    def _create_mitigation_sheet(self, writer, results: List[Dict]):
        """Create mitigation actions sheet"""
        import pandas as pd

        mitigation_data = []

        for result in results:
            breach_info = result.get('breach_info') or {}
            if breach_info.get('found') and breach_info.get('mitigation_steps'):
                for i, step in enumerate(breach_info['mitigation_steps'], 1):
                    mitigation_data.append({
                        'Email': result.get('email', 'N/A'),
                        'Priority': i,
                        'Action Required': step,
                        'Severity': (breach_info.get('severity') or 'N/A').upper(),
                        'Status': 'Pending',
                        'Assigned To': '',
                        'Due Date': '',
                        'Notes': ''
                    })

        if mitigation_data:
            df_mitigation = pd.DataFrame(mitigation_data)
            df_mitigation.to_excel(writer, sheet_name='Mitigation Actions', index=False)
        else:
            # Create placeholder row
            empty_data = pd.DataFrame([{
                'Email': 'No mitigation actions required',
                'Priority': '-',
                'Action Required': 'All emails are secure',
                'Status': 'N/A'
            }])
            empty_data.to_excel(writer, sheet_name='Mitigation Actions', index=False)

    def _create_dns_security_sheet(self, writer, results: List[Dict]):
        """Create DNS security analysis sheet"""
        import pandas as pd

        dns_data = []

        for result in results:
            dns = result.get('dns_security') or {}
            if dns:
                dns_data.append({
                    'Domain': result.get('email', 'N/A').split('@')[1] if '@' in result.get('email', '') else 'N/A',
                    'Email': result.get('email', 'N/A'),
                    'DNS Score': dns.get('score', 0),
                    'SPF Record': 'Configured' if dns.get('spf') else 'Missing',
                    'DMARC Policy': 'Configured' if dns.get('dmarc') else 'Missing',
                    'DKIM': 'Configured' if dns.get('dkim') else 'Missing',
                    'DNSSEC': 'Enabled' if dns.get('dnssec') else 'Disabled',
                    'MX Records': 'Found' if dns.get('mx') else 'Missing',
                    'BIMI': 'Yes' if dns.get('bimi') else 'No',
                    'MTA-STS': 'Yes' if dns.get('mta_sts') else 'No',
                    'TLS-RPT': 'Yes' if dns.get('tls_rpt') else 'No',
                    'Issues': ', '.join(str(i) for i in dns.get('issues', [])) if dns.get('issues') else 'None'
                })

        if dns_data:
            df_dns = pd.DataFrame(dns_data)
            # Remove duplicates by domain
            df_dns = df_dns.drop_duplicates(subset=['Domain'])
            df_dns.to_excel(writer, sheet_name='DNS Security', index=False)
        else:
            # Create placeholder row
            empty_data = pd.DataFrame([{
                'Domain': 'No DNS data available',
                'SPF Record': '-',
                'DMARC Policy': '-',
                'DKIM': '-',
                'DNSSEC': '-'
            }])
            empty_data.to_excel(writer, sheet_name='DNS Security', index=False)

    def _create_threat_details_sheet(self, writer, results: List[Dict]):
        """Create per-email threat breakdown sheet with all detection details"""
        import pandas as pd

        threat_data = []

        for result in results:
            email = result.get('email', 'N/A')
            for threat in (result.get('threats') or []):
                threat_data.append({
                    'Email': email,
                    'Threat Type': threat.get('type', 'unknown').replace('_', ' ').title(),
                    'Description': threat.get('description', 'N/A'),
                    'Severity': (threat.get('severity') or 'N/A').upper(),
                    'Confidence': f"{threat.get('confidence', 0):.0%}" if threat.get('confidence') else 'N/A',
                    'Risk Score': result.get('risk_score', 0),
                    'Risk Level': (result.get('risk_level') or 'unknown').upper()
                })

        if threat_data:
            df = pd.DataFrame(threat_data)
            # Sort by severity (critical first) then risk score
            severity_order = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3, 'N/A': 4}
            df['_sort'] = df['Severity'].map(severity_order).fillna(4)
            df = df.sort_values(['_sort', 'Risk Score'], ascending=[True, False]).drop('_sort', axis=1)
            df.to_excel(writer, sheet_name='Threat Details', index=False)
        else:
            empty_data = pd.DataFrame([{
                'Email': 'No threats detected',
                'Threat Type': '-',
                'Description': 'All analyzed emails appear safe',
                'Severity': '-'
            }])
            empty_data.to_excel(writer, sheet_name='Threat Details', index=False)

    def _create_disposable_typosquat_sheet(self, writer, results: List[Dict]):
        """Create disposable email & typosquatting detection sheet"""
        import pandas as pd

        detection_data = []

        for result in results:
            email = result.get('email', 'N/A')
            threats = result.get('threats') or []
            typosquat_info = result.get('typosquat_info') or {}

            # Check for disposable
            is_disposable = any(t.get('type') == 'disposable_email' for t in threats)
            disp_confidence = next((t.get('confidence', 0) for t in threats if t.get('type') == 'disposable_email'), 0)

            # Check for typosquatting
            is_typosquat = typosquat_info.get('is_typosquat', False)

            if is_disposable or is_typosquat:
                detection_data.append({
                    'Email': email,
                    'Disposable Email': 'YES' if is_disposable else 'NO',
                    'Disposable Confidence': f'{disp_confidence:.0%}' if is_disposable else 'N/A',
                    'Typosquatting': 'YES' if is_typosquat else 'NO',
                    'Impersonates': typosquat_info.get('target_domain', 'N/A') if is_typosquat else 'N/A',
                    'Attack Type': (typosquat_info.get('attack_type') or 'N/A').replace('_', ' ').title() if is_typosquat else 'N/A',
                    'Similarity': f"{typosquat_info.get('similarity', 0):.0%}" if is_typosquat else 'N/A',
                    'Risk Score': result.get('risk_score', 0),
                    'Risk Level': (result.get('risk_level') or 'unknown').upper()
                })

        if detection_data:
            df = pd.DataFrame(detection_data)
            df = df.sort_values('Risk Score', ascending=False)
            df.to_excel(writer, sheet_name='Fraud Detection', index=False)
        else:
            empty_data = pd.DataFrame([{
                'Email': 'No disposable or typosquatting emails detected',
                'Status': 'Clean',
                'Note': 'All email domains appear legitimate'
            }])
            empty_data.to_excel(writer, sheet_name='Fraud Detection', index=False)

    def _create_vulnerability_sheet(self, writer, results: List[Dict]):
        """Create vulnerability assessment sheet"""
        import pandas as pd

        vuln_data = []

        for result in results:
            email = result.get('email', 'N/A')
            domain = email.split('@')[1] if '@' in email else 'N/A'

            for vuln in result.get('vulnerabilities', []):
                vuln_data.append({
                    'Domain': domain,
                    'Email': email,
                    'Vulnerability Type': vuln.get('type', 'unknown').replace('_', ' ').title(),
                    'Description': vuln.get('description', 'N/A'),
                    'Severity': (vuln.get('severity') or 'N/A').upper(),
                    'Remediation': vuln.get('remediation', 'N/A'),
                    'Status': 'Open',
                    'Priority': '',
                    'Assigned To': '',
                    'Notes': ''
                })

        if vuln_data:
            df = pd.DataFrame(vuln_data)
            # Sort by severity
            severity_order = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3, 'N/A': 4}
            df['_sort'] = df['Severity'].map(severity_order).fillna(4)
            df = df.sort_values('_sort').drop('_sort', axis=1)
            # De-duplicate by domain + vulnerability type
            df = df.drop_duplicates(subset=['Domain', 'Vulnerability Type'])
            df.to_excel(writer, sheet_name='Vulnerability Assessment', index=False)
        else:
            empty_data = pd.DataFrame([{
                'Domain': 'No vulnerabilities detected',
                'Vulnerability Type': '-',
                'Description': 'All domains passed security checks',
                'Severity': '-'
            }])
            empty_data.to_excel(writer, sheet_name='Vulnerability Assessment', index=False)

    def _create_advanced_security_sheet(self, writer, results: List[Dict]):
        """Create advanced security checks sheet (DNSBL, CT, DGA, ThreatFox, Gravatar, Parked)"""
        import pandas as pd

        adv_data = []
        for result in results:
            email = result.get('email', 'N/A')
            domain = email.split('@')[1] if '@' in email else 'N/A'
            dnsbl = result.get('dnsbl') or {}
            ct = result.get('cert_transparency') or {}
            dga = result.get('dga_analysis') or {}
            tfox = result.get('threatfox') or {}
            grav = result.get('gravatar') or {}
            park = result.get('parked_domain') or {}

            adv_data.append({
                'Email': email,
                'Domain': domain,
                'DNSBL Status': f'Listed ({dnsbl.get("listed_count", 0)})' if dnsbl.get('listed') else 'Clean',
                'DNSBL IPs Checked': ', '.join(str(ip) for ip in dnsbl.get('checked_ips', [])) if dnsbl.get('checked_ips') else 'N/A',
                'ThreatFox IOCs': tfox.get('ioc_count', 0) if tfox.get('found') else 0,
                'ThreatFox Status': 'IOC Found' if tfox.get('found') else 'Clean',
                'DGA Score': round(float(dga.get('dga_score') or 0.0), 3),
                'DGA Detected': 'Yes' if dga.get('is_dga') else 'No',
                'Cert Count': ct.get('cert_count', 0) if ct.get('found') else 0,
                'Cert First Seen': ct.get('first_seen', 'N/A') or 'N/A',
                'Cert Last Seen': ct.get('last_seen', 'N/A') or 'N/A',
                'Gravatar Profile': 'Yes' if grav.get('has_profile') else 'No',
                'Parked Domain': 'Yes' if park.get('is_parked') else 'No',
                'Parking Indicators': ', '.join(str(i) for i in park.get('indicators', [])[:3]) if park.get('indicators') else 'None',
            })

        if adv_data:
            df = pd.DataFrame(adv_data)
            df.to_excel(writer, sheet_name='Advanced Security', index=False)
        else:
            empty_data = pd.DataFrame([{
                'Email': 'No advanced security data',
                'DNSBL Status': '-',
                'ThreatFox Status': '-',
                'DGA Detected': '-',
            }])
            empty_data.to_excel(writer, sheet_name='Advanced Security', index=False)

    def _create_data_exposure_sheet(self, writer, results: List[Dict]):
        """Create data exposure summary sheet"""
        import pandas as pd
        from collections import Counter

        # Aggregate data classes
        data_class_counts = Counter()
        breach_timeline = []
        password_breaches = []

        for result in results:
            email = result.get('email', 'N/A')
            breach_info = result.get('breach_info') or {}

            if breach_info.get('found'):
                for breach in (breach_info.get('details') or []):
                    if isinstance(breach, dict):
                        data_classes = breach.get('data_classes') or []
                        for dc in data_classes:
                            if isinstance(dc, str):
                                data_class_counts[dc] += 1

                        breach_timeline.append({
                            'Breach Name': breach.get('name', 'Unknown'),
                            'Breach Date': breach.get('breach_date', 'Unknown'),
                            'Affected Email': email,
                            'Affected Accounts': breach.get('pwn_count', 'N/A'),
                            'Data Types Exposed': ', '.join(str(dc) for dc in (breach.get('data_classes') or [])[:8]),
                            'Domain': breach.get('domain', 'N/A')
                        })

            # Password breach
            pwd_breach = result.get('password_breach') or {}
            if pwd_breach.get('found'):
                password_breaches.append({
                    'Email': email,
                    'Password Status': 'COMPROMISED',
                    'Action Required': 'CHANGE IMMEDIATELY',
                    'Risk Score': result.get('risk_score', 0)
                })

        # Sheet 1: Data Classes Summary
        if data_class_counts:
            high_sensitivity = ['Passwords', 'Credit cards', 'Bank account numbers', 'Social security numbers',
                                'Credit card CVV', 'PINs', 'Security questions and answers', 'Auth tokens']

            exposure_data = []
            for dc, count in data_class_counts.most_common():
                exposure_data.append({
                    'Data Type': dc,
                    'Exposure Count': count,
                    'Sensitivity': 'HIGH' if dc in high_sensitivity else 'MEDIUM',
                    'Risk Impact': 'Critical - Immediate action needed' if dc in high_sensitivity else 'Moderate - Monitor closely'
                })

            df_exposure = pd.DataFrame(exposure_data)
            df_exposure.to_excel(writer, sheet_name='Data Exposure', index=False)

            # Add breach timeline below
            if breach_timeline:
                df_timeline = pd.DataFrame(breach_timeline)
                df_timeline.to_excel(writer, sheet_name='Data Exposure', index=False,
                                     startrow=len(exposure_data) + 4)

            # Add password breaches if any
            if password_breaches:
                df_pwd = pd.DataFrame(password_breaches)
                offset = len(exposure_data) + 4 + (len(breach_timeline) + 3 if breach_timeline else 0)
                df_pwd.to_excel(writer, sheet_name='Data Exposure', index=False, startrow=offset)
        else:
            empty_data = pd.DataFrame([{
                'Data Type': 'No data exposure detected',
                'Exposure Count': 0,
                'Sensitivity': '-',
                'Note': 'No emails found in breach databases'
            }])
            empty_data.to_excel(writer, sheet_name='Data Exposure', index=False)

    def _create_ml_metrics_sheet(self, writer):
        """Create ML Prediction Statistics sheet from ACTUAL email analysis"""
        import pandas as pd

        try:
            # Use the REAL analyzer instance passed during initialization
            if not self.analyzer or not hasattr(self.analyzer, 'ml_engine'):
                # No analyzer provided - create informational placeholder
                empty_data = pd.DataFrame([{
                    'Information': 'ML Prediction Statistics will appear here after email analysis',
                    'Note': 'Analyze emails with ML enabled to see prediction statistics'
                }])
                empty_data.to_excel(writer, sheet_name='ML Prediction Statistics', index=False)
                logger.info("ML sheet created with placeholder (no analyzer)")
                return

            ml_engine = self.analyzer.ml_engine

            # Check if any predictions have been made
            if not hasattr(ml_engine, 'prediction_history') or not ml_engine.prediction_history:
                # No predictions made yet
                empty_data = pd.DataFrame([{
                    'Information': 'No email predictions have been made yet',
                    'Note': 'ML prediction statistics will appear here after you analyze emails',
                    'Status': 'Waiting for email analysis...'
                }])
                empty_data.to_excel(writer, sheet_name='ML Prediction Statistics', index=False)
                logger.info("ML sheet created - no predictions yet")
                return

            # Calculate prediction statistics from actual predictions
            prediction_stats = self._calculate_prediction_statistics(ml_engine.prediction_history)

            # Create prediction statistics sheet
            stats_data = [{
                'Metric': 'Total Emails Analyzed',
                'Value': len(ml_engine.prediction_history),
                'Description': 'Number of emails analyzed with ML models'
            }]

            # Add model-specific prediction counts
            for model_name in ml_engine.models.keys():
                predictions_for_model = [p for p in ml_engine.prediction_history
                                        if model_name in p.get('predictions', {})]
                if predictions_for_model:
                    avg_confidence = sum((p['predictions'].get(model_name) or 0)
                                        for p in predictions_for_model) / len(predictions_for_model)
                    stats_data.append({
                        'Metric': f'{model_name.replace("_", " ").title()} - Avg Confidence',
                        'Value': f'{avg_confidence:.1%}',
                        'Description': f'Average prediction confidence for {model_name}'
                    })

            df_stats = pd.DataFrame(stats_data)
            df_stats.to_excel(writer, sheet_name='ML Prediction Statistics', index=False)

            logger.info(f"Created ML prediction statistics from {len(ml_engine.prediction_history)} actual predictions")

        except Exception as e:
            logger.warning(f"Could not create ML prediction statistics: {e}")
            # Create informational placeholder on error
            try:
                empty_data = pd.DataFrame([{
                    'Information': 'ML prediction statistics unavailable',
                    'Reason': f'{str(e)[:100]}'
                }])
                empty_data.to_excel(writer, sheet_name='ML Prediction Statistics', index=False)
            except Exception:
                pass

    def _calculate_prediction_statistics(self, prediction_history: list) -> dict:
        """Calculate statistics from actual predictions"""
        if not prediction_history:
            return {}

        stats = {
            'total_predictions': len(prediction_history),
            'model_predictions': {}
        }

        # Calculate per-model statistics
        for pred in prediction_history:
            for model_name, confidence in pred.get('predictions', {}).items():
                if model_name not in stats['model_predictions']:
                    stats['model_predictions'][model_name] = []
                stats['model_predictions'][model_name].append(confidence)

        return stats

    def _create_ml_charts_sheet(self, writer, results: List[Dict]):
        """Create ML Performance Charts sheet with visualizations"""
        import pandas as pd

        try:
            from openpyxl.chart import BarChart, PieChart, Reference
            from openpyxl.chart.label import DataLabelList

            # Create data for ML model performance comparison
            ml_data = []

            # Keys that are metadata, not model predictions
            skip_keys = {'ensemble', 'is_malicious', 'precision_threshold', 'anomaly_score'}

            for result in results:
                if result.get('ml_predictions') and self.config.enable_ml:
                    for model, score in result['ml_predictions'].items():
                        if model in skip_keys:
                            continue
                        if not isinstance(score, (int, float)):
                            continue
                        ml_data.append({
                            'Model': model.replace('_', ' ').title(),
                            'Score': float(score) * 100
                        })

            if not ml_data:
                # No ML data available
                empty_data = pd.DataFrame([{
                    'Information': 'No ML prediction data available for charts',
                    'Note': 'Analyze emails with ML enabled to see visualizations'
                }])
                empty_data.to_excel(writer, sheet_name='ML Performance Charts', index=False)
                logger.info("ML charts sheet created with placeholder")
                return

            # Calculate average scores per model
            df = pd.DataFrame(ml_data)
            avg_scores = df.groupby('Model')['Score'].mean().reset_index()
            avg_scores.columns = ['Model Name', 'Average Prediction Score (%)']

            # Write data to Excel
            avg_scores.to_excel(writer, sheet_name='ML Performance Charts', index=False, startrow=1)

            # Get workbook to add charts
            workbook = writer.book
            worksheet = writer.sheets['ML Performance Charts']

            # Add title using openpyxl API
            from openpyxl.styles import Font
            title_cell = worksheet.cell(row=1, column=1, value='ML MODEL PERFORMANCE COMPARISON')
            title_cell.font = Font(bold=True, size=14)

            # Chart 1: Bar Chart - Model Performance Comparison
            try:
                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                chart.title = "Average ML Model Predictions"
                chart.y_axis.title = 'Average Score (%)'
                chart.x_axis.title = 'Model'

                # Reference data
                data = Reference(worksheet,
                               min_col=2, min_row=2,
                               max_row=len(avg_scores) + 2)
                cats = Reference(worksheet,
                               min_col=1, min_row=3,
                               max_row=len(avg_scores) + 2)

                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.height = 10
                chart.width = 20

                worksheet.add_chart(chart, "E2")
                logger.info("Added ML bar chart to Excel")
            except Exception as e:
                logger.warning(f"Could not create ML bar chart: {e}")

            # Add confidence distribution data
            confidence_ranges = {
                'High (>80%)': sum(1 for d in ml_data if d['Score'] > 80),
                'Medium (50-80%)': sum(1 for d in ml_data if 50 <= d['Score'] <= 80),
                'Low (<50%)': sum(1 for d in ml_data if d['Score'] < 50)
            }

            conf_df = pd.DataFrame([
                {'Confidence Level': k, 'Count': v}
                for k, v in confidence_ranges.items()
            ])
            conf_df.to_excel(writer, sheet_name='ML Performance Charts',
                            index=False, startrow=len(avg_scores) + 5)

            # Chart 2: Pie Chart - Confidence Distribution
            try:
                pie = PieChart()
                pie.title = "ML Prediction Confidence Distribution"

                labels = Reference(worksheet,
                                 min_col=1,
                                 min_row=len(avg_scores) + 7,
                                 max_row=len(avg_scores) + 9)
                data = Reference(worksheet,
                               min_col=2,
                               min_row=len(avg_scores) + 6,
                               max_row=len(avg_scores) + 9)

                pie.add_data(data, titles_from_data=True)
                pie.set_categories(labels)
                pie.height = 10
                pie.width = 12

                worksheet.add_chart(pie, f"E{len(avg_scores) + 6}")
                logger.info("Added ML pie chart to Excel")
            except Exception as e:
                logger.warning(f"Could not create ML pie chart: {e}")

            logger.info(f"Created ML charts sheet with {len(avg_scores)} models")

        except Exception as e:
            logger.warning(f"Could not create ML charts sheet: {e}")
            try:
                empty_data = pd.DataFrame([{
                    'Information': 'ML charts unavailable',
                    'Reason': f'{str(e)[:100]}'
                }])
                empty_data.to_excel(writer, sheet_name='ML Performance Charts', index=False)
            except Exception:
                pass

    def _apply_advanced_formatting(self, filename: str, stats: Dict):
        """Apply advanced formatting to the Excel workbook"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import PieChart, BarChart, Reference
            from openpyxl.utils import get_column_letter

            wb = load_workbook(filename)

            # CRITICAL: Ensure at least one sheet is visible
            # Excel requires at least one visible sheet
            visible_sheets = [sheet for sheet in wb.worksheets if sheet.sheet_state == 'visible']
            if len(visible_sheets) == 0:
                # Make the first sheet visible
                if len(wb.worksheets) > 0:
                    wb.worksheets[0].sheet_state = 'visible'
                    logger.info("Forced first sheet to be visible")

            # Define styles
            header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1E3C72', end_color='1E3C72', fill_type='solid')
            title_font = Font(name='Calibri', size=16, bold=True, color='1E3C72')

            border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )

            # Format each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Ensure sheet is visible
                if hasattr(ws, 'sheet_state'):
                    ws.sheet_state = 'visible'

                # Format headers (row 1) only if there are rows
                if ws.max_row > 0 and ws.max_column > 0:
                    for cell in ws[1]:
                        if cell.value:  # Only format cells with content
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = border

                    # Auto-adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter

                        for cell in column:
                            try:
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except Exception:
                                pass

                        adjusted_width = min(max_length + 4, 60)
                        ws.column_dimensions[column_letter].width = adjusted_width

                    # Apply conditional formatting for risk levels
                    if sheet_name in ['Detailed Results', 'Risk Analysis']:
                        self._apply_conditional_formatting(ws, sheet_name)

            # Add chart to Risk Analysis sheet
            if 'Risk Analysis' in wb.sheetnames:
                try:
                    self._add_risk_chart(wb['Risk Analysis'])
                except Exception as e:
                    logger.warning(f"Could not add chart to Risk Analysis: {e}")

            # Final check: ensure at least one sheet is visible before saving
            visible_count = sum(1 for sheet in wb.worksheets if sheet.sheet_state == 'visible')
            if visible_count == 0:
                wb.worksheets[0].sheet_state = 'visible'
                logger.warning("No visible sheets detected, forced first sheet visible")

            # Save formatted workbook
            wb.save(filename)
            logger.info(f"Successfully formatted workbook with {len(wb.worksheets)} sheets, {visible_count} visible")

        except Exception as e:
            logger.error(f"Error applying formatting: {e}")
            # Don't re-raise - we want the basic Excel file to still be created

    def _apply_conditional_formatting(self, ws, sheet_name: str):
        """Apply conditional formatting for risk levels"""
        from openpyxl.styles import PatternFill

        # Define color fills
        critical_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        high_fill = PatternFill(start_color='FFE5CC', end_color='FFE5CC', fill_type='solid')
        medium_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        low_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')

        # Find Risk Level column dynamically by header name
        risk_col = None
        for cell in ws[1]:
            if cell.value and 'risk' in str(cell.value).lower() and 'level' in str(cell.value).lower():
                risk_col = cell.column
                break

        if risk_col:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    if cell.column == risk_col:
                        value = str(cell.value).upper() if cell.value else ''
                        if 'CRITICAL' in value:
                            cell.fill = critical_fill
                        elif 'HIGH' in value:
                            cell.fill = high_fill
                        elif 'MEDIUM' in value:
                            cell.fill = medium_fill
                        elif 'LOW' in value or 'MINIMAL' in value:
                            cell.fill = low_fill

    def _add_risk_chart(self, ws):
        """Add pie chart for risk distribution"""
        try:
            from openpyxl.chart import PieChart, Reference

            # Only add chart if there's enough data
            if ws.max_row < 2:
                logger.info("Skipping chart creation - not enough data rows")
                return

            pie = PieChart()
            labels = Reference(ws, min_col=1, min_row=2, max_row=5)
            data = Reference(ws, min_col=2, min_row=1, max_row=5)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "Risk Level Distribution"
            pie.height = 10
            pie.width = 15

            ws.add_chart(pie, "F2")
            logger.info("Risk chart added successfully")

        except Exception as e:
            logger.warning(f"Could not add chart (non-critical): {e}")

    def _calculate_stats(self, results: List[Dict]) -> Dict:
        """Calculate comprehensive statistics including all detection types"""
        total = len(results)
        critical = sum(1 for r in results if r.get('risk_level') == 'critical')
        high = sum(1 for r in results if r.get('risk_level') == 'high')
        medium = sum(1 for r in results if r.get('risk_level') == 'medium')
        low = sum(1 for r in results if r.get('risk_level') in ['low', 'minimal'])
        breached = sum(1 for r in results if (r.get('breach_info') or {}).get('found'))

        total_risk = sum(r.get('risk_score', 0) for r in results)
        avg_risk = int(total_risk / total) if total > 0 else 0
        max_risk = max((r.get('risk_score', 0) for r in results), default=0)

        breach_percentage = (breached / total * 100) if total > 0 else 0

        # DNS issues
        dns_issues = sum(1 for r in results if (r.get('dns_security') or {}).get('issues'))

        # Total threats
        total_threats = sum(len(r.get('threats', [])) for r in results)

        # New detection counts
        disposable_count = sum(1 for r in results
                               if any(t.get('type') == 'disposable_email' for t in r.get('threats', [])))
        typosquat_count = sum(1 for r in results
                              if (r.get('typosquat_info') or {}).get('is_typosquat'))
        password_breach_count = sum(1 for r in results
                                    if (r.get('password_breach') or {}).get('found'))
        vulnerability_count = sum(len(r.get('vulnerabilities') or []) for r in results)

        return {
            'total': total,
            'critical': critical,
            'high': high,
            'medium': medium,
            'low': low,
            'breached': breached,
            'avg_risk': avg_risk,
            'max_risk': max_risk,
            'breach_percentage': breach_percentage,
            'critical_percentage': (critical / total * 100) if total > 0 else 0,
            'high_percentage': (high / total * 100) if total > 0 else 0,
            'medium_percentage': (medium / total * 100) if total > 0 else 0,
            'low_percentage': (low / total * 100) if total > 0 else 0,
            'dns_issues': dns_issues,
            'total_threats': total_threats,
            'disposable_emails': disposable_count,
            'typosquatting_domains': typosquat_count,
            'password_breaches': password_breach_count,
            'vulnerabilities': vulnerability_count
        }
